import time
import logging
import fitz
import pandas as pd
from .tag_keys import SKIP_KEYS, BEGIN_KEYS, INFO_KEYS, DATE_PATTERN, MONTH, PROTOCOLS, PATH,BasicDict
from .analyze_class import check_dict, keyword_modify, keyword_contains_in, resove_dict
import string
from datetime import datetime
from operator import itemgetter
from itertools import groupby
import re
from .mapper import predefine_, pdf_df_rename, get_protocol, Mapper
import os
import pdfminer
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFTextExtractionNotAllowed
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.layout import LAParams, LTChar, LTCurve, LTTextLine, LTText, LTTextBox, LTLine, LTRect, LTImage, \
    LTTextBoxHorizontal, LTPage
from pdfminer.converter import PDFPageAggregator
from pdfminer.pdfpage import PDFPage
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
# local

from .pdf_layouts import Point, Cluster, Cell, Table, approxiamtion, inside
from .Document_simulation import Document

from .main_decorators import timeout, check_first_page_with_dict, check_not_first_page_with_dict, controller


LINE_MARGIN = 0.09
CHAR_MARGIN = 0.5
WORD_MARGIN = 0.5


@controller()
def user_select_from_list(name=None, item=None):
    """
    change to online for as a spindbox let user choose which protocol to use if not locate one"""
    print("select %s as %s" % (item, name))
    return item


def checkpointsvaliadation(point_list, textboxlist):
    # remove all points in a textbox e.g a html address with bottom line
    valid_points = []
    unvalid_points = []
    for point in point_list:
        unvalid = False
        for t in textboxlist:
            if all([point.x >= t.bbox[0], point.x <= t.bbox[2],
                    point.y >= t.bbox[1], point.y <= t.bbox[3]]):
                unvalid = True
        if unvalid:
            unvalid_points.append(point)
        else:
            valid_points.append(point)

    return valid_points


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """    

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def find_key(*args, **kwargs):
    element = kwargs['element']
    if isinstance(element, LTTextBox):
        current = keyword_modify(element.get_text())
        done = check_dict(
            item=current, the_dict=kwargs['the_dict'], *args, start=kwargs['start'])
        for do in done:
            if do[2]:
                return do, element
    return

def find_right_neightbors(media_bbox, layout, obj, width=3, pts=0.2, position='bbox'):
    """
    top mode is used when information is located on the right of the keyword, but
    possibly with a different height, also used for cell obj
    :param media_bbox:
    :param layout:
    :param obj:
    :param width:
    :param pts:
    :param position:
    :return:
    """

    class area:
        def __init__(self, bbox):
            self.bbox = bbox

    if position == 'bbox':
        target_bbox = (obj.bbox[2], obj.bbox[1], media_bbox[2], obj.bbox[3])
        current = area(target_bbox)
        for element in layout:
            if isinstance(element, LTTextBox):
                if element.bbox[2] - element.bbox[0] > 3:
                    if inside(bigger_obj=current, smaller_obj=element, width=width, pts=pts):
                        yield keyword_modify(element.get_text(), special=True)
    elif position == 'top':
        alignment = int(obj.bbox[3])
        _right_neightbors = []
        for item in layout:
            if isinstance(item, LTTextBox):
                if approxiamtion(item.bbox[3], alignment) and item.bbox[0] > obj.bbox[0]:
                    logging.info("this is item : %s" % item)
                    _right_neightbors.append(item)
        _right_neightbors.sort(key=lambda x: x.bbox[0])

        if _right_neightbors:
            j = _right_neightbors.pop(0).get_text()
            if len(j.strip()) > 3:
                # not null 
                yield j
        else:
            find_right_neightbors(media_bbox == media_bbox, layout=layout, obj=obj)



class PageExtractor:

    def __init__(self, layout, pageid,page_number, cache=True):
        self.layout = layout
        self.pageid = pageid
        self.page_number = page_number
        self.text_box = []
        self.cache = cache
        self.body = None
        self.tables = []
        

    def sort_text_box(self):
        # from top to bottom, left to right
        self.text_box = sorted(sorted(self.text_box, key=lambda x: x.bbox[0]),
                               key=lambda x: round(x.bbox[3]), reverse=True)

    def sort_tables(self):
        # by y1
        if self.tables:
            self.tables = sorted(self.tables, key=lambda x: x.bbox[3], reverse=True)

    def __repr__(self):
        return "<pageid: %s, total tables: %s >" % (self.pageid, len(self.tables))

    def __call__(self, *args, **kwargs):
        self.solve_points()
        self.cluster_to_cell()
        self.cell_to_table()

    def solve_points(self):
        Point.set_page_points(pageid=self.pageid)
        point_count = 0
        mid_count = 0
        temp = set()
        # lines = []
        for element in self.layout:
            logging.debug("checking element %s" % element)
            try:
                if element._objs:
                    for ele in element._objs:
                        temp.update([ele])
                else:
                    logging.debug("single element : %s" % element)
                    temp.update([element])
            except AttributeError:
                temp.update([element])

        for element in temp:
            if isinstance(element, LTPage):
                logging.debug("this is ltpage %s" % element)
            if isinstance(element, LTCurve):
                point_left_bottom = Point(x=element.bbox[0], y=element.bbox[1], pageid=self.pageid,
                                          point_id=point_count)
                point_right_top = Point(element.bbox[2], element.bbox[3], pageid=self.pageid, point_id=point_count + 1)
                point_left_top = Point(x=element.bbox[0], y=element.bbox[3], pageid=self.pageid,
                                       point_id=point_count + 2)
                point_right_bottom = Point(element.bbox[2], element.bbox[1], pageid=self.pageid,
                                           point_id=point_count + 3)
                middle_point_1 = Point(x=(element.bbox[0] + element.bbox[2]) / 2,
                                       y=element.bbox[1], pageid=self.pageid, point_id=mid_count,
                                       is_child=True)
                middle_point_2 = Point(x=element.bbox[2],
                                       y=(element.bbox[1] + element.bbox[3]) / 2, pageid=self.pageid,
                                       point_id=mid_count + 1,
                                       is_child=True)
                middle_point_3 = Point(x=(element.bbox[0] + element.bbox[2]) / 2,
                                       y=element.bbox[3], pageid=self.pageid,
                                       point_id=mid_count + 2,
                                       is_child=True)
                middle_point_4 = Point(x=element.bbox[0],
                                       y=(element.bbox[1] + element.bbox[3]) / 2, pageid=self.pageid,
                                       point_id=mid_count + 3,
                                       is_child=True)
                point_left_bottom.add_to_page_points(pageid=self.pageid)
                point_right_top.add_to_page_points(pageid=self.pageid)
                point_right_bottom.add_to_page_points(pageid=self.pageid)
                point_left_top.add_to_page_points(pageid=self.pageid)
                point_left_bottom.assign_child(middle_point_1)
                point_left_bottom.assign_child(middle_point_4)
                point_left_top.assign_child(middle_point_4)
                point_left_top.assign_child(middle_point_3)
                point_right_top.assign_child(middle_point_2)
                point_right_top.assign_child(middle_point_3)
                point_right_bottom.assign_child(middle_point_1)
                point_right_bottom.assign_child(middle_point_2)
                middle_point_1.add_to_page_points(pageid=self.pageid)
                middle_point_2.add_to_page_points(pageid=self.pageid)
                middle_point_3.add_to_page_points(pageid=self.pageid)
                middle_point_4.add_to_page_points(pageid=self.pageid)
                point_count += 4
                mid_count += 4
            elif isinstance(element, (LTTextBox, LTTextLine)):
                logging.info("adding textbox %s", element)
                self.text_box.append(element)

        assert len(self.text_box) != 0, "text box length = %d" % len(self.text_box)
        self.sort_text_box()

    def cluster_to_cell(self):
        clean_points = Point.pagePoints[self.pageid]
        if len(clean_points) == 0:
            logging.warning("clean points is empty")
        Cluster.group_cluster(pageid=self.pageid, cleaned_points=clean_points)
        Cluster.sort_page_cluster(pageid=self.pageid)
        Cluster.assign_id(pageid=self.pageid)
        for c in Cluster.pageSortedClusters[self.pageid]:
            if len(c.value) < 3:
                logging.warning("%s cluster has less than 3 children", c)
        for c in Cluster.pageSortedClusters[self.pageid]:
            c.get_middle()
        Cluster.locate_roots(pageid=self.pageid)
        Cluster.do_cell(pageid=self.pageid)

        for c in list(Cell.pageCells[self.pageid].values()):
            for j in c:
                j()
        Cell.assign_text(pageid=self.pageid, textlist=self.text_box)
        Cell.check_text(self.pageid)

    def cell_to_table(self):
        table_generator = Table.get_tables(list(Cell.pageCells[self.pageid].values()), pageid=self.pageid)

        for table in table_generator:
            self.tables.append(next(table))

    def get_body(self, idx=-1):
        """
        transform a table obj into pandas DataFrane, according to the pdf structure, set idx default value == -1
        :return: last table object in a page
        """

        if self.tables:
            self.sort_tables()
            self.body = self.tables[idx].table_to_df()
        return self.body

    def page_to_excel(self, ws, start_row=2, start_col=2, ref=None):
        start_row = start_row
        start_col = start_col
        if self.tables:
            self.sort_tables()
            for t in self.tables:
                t.table_to_excel(worksheet=ws, start_row=start_row, start_col=start_col, ref=ref)
                start_row += t.base_matrix.shape[0] + 1
        if len(self.text_box) > 0:
            letter = get_column_letter(start_col + 10)
            for i in range(len(self.text_box)):
                ws[letter  + str(i + 1)] = self.text_box[i].get_text()

    def inside_table_detect(self, key, table):
        """ UNDONE
        find keywords from table
        :param key:
        :param table:
        :return:
        """
       
        key = key
       
        with pd.option_context('display.max_rows', None, 'display.max_columns', None):
            table.insert(loc=0,
              value=table[table.columns[0]].apply(lambda x: 
              re.sub('[\n\t]', '', ' '.join(keyword_modify(x,special=True).lower().split()))),
              column='reitem')
        
class FirstPageInfo:
        """
        collect information from layouts in first page, mainly in first page
        atttr = {
            Vendor Name : vendor name : str => done
            Factory Name: factory name: str => done
            Report issue date : report_issue_date : str => guessing format not done
            Protocol used for review: protocol : str => done
            All reviewed  Report  number :  report number : str => done
            Lab: lab: tuple(lab name, region) can related to dict => done
        }
        input :
            layout : pdfminer.ltpage
            attr : assign an external dict
            media_bbox: media bbox of first page
        function:
            get_info() : main function, update items to attr
            get_report_issue_date(): return largest date according to regex and given date format,returns a date type
            refine_lab(): regex to refine lab info
            refine_vendor: regex to refine vendor name && factory name
        output:
            a dict contains import information to back fill to tracker,
            if any part not located, should returns a warning
        """

        def __init__(self, page, layout, attr, media_bbox):
            self.page = page
            self.layout = layout
            self.attr = attr
            self.media_bbox = media_bbox
            self.potential_date_list = set()
            self.first_keys = ['report_number', 'protocol', 'vendor_name', 'factory_name', 'lab']
            if not isinstance(self.attr, dict):
                raise TypeError

        def __call__(self, *args, **kwargs):
            self.get_info()
            self.get_report_issue_date(pattern='oo/dd/yyyy')
            self.refine('report_number')
            self.refine_lab()
            self.refine_vender()

        def get_info(self):
            controler = [False] * len(self.first_keys)
            slot = [None] * len(self.first_keys)
            for element in self.layout:
                if all(controler):
                    break
                else:
                    if isinstance(element, LTTextBox):
                        if element.bbox[2] - element.bbox[0] > 3:
                            print(element)
                            for i in range(len(controler)):
                                if not controler[i]:
                                    slot[i] = find_key(self.first_keys[i], element=element, the_dict=INFO_KEYS,
                                                    start=0)
                                    # print("slot[{}] is {}".format(i, slot[i]))
                                    if slot[i]:
                                        content = find_right_neightbors(self.media_bbox, self.layout, slot[i][1],
                                                                        position='top')
                                        try:
                                            t = next(content)
                                            self.attr[self.first_keys[i]] = t.strip()
                                        except StopIteration:
                                            content1 = find_right_neightbors(self.media_bbox, self.layout, slot[i][1],
                                                                            width=5, pts=0.5)
                                            t1 = next(content1)
                                            self.attr[self.first_keys[i]] = t1
                                        controler[i] = True
                                    else:
                                        try:
                                            current_dict = resove_dict(self.first_keys[i], the_dict=INFO_KEYS,
                                                                    start=0)
                                            # print("current_dict is ", current_dict)
                                            for k in current_dict:
                                                # print("this is k", k, element.get_text())
                                                t = keyword_contains_in(keyword=k, target_str=element.get_text(),
                                                                        split_mark=":", position=2, title=True)
                                                # print("this is t", t)
                                                if t:
                                                    self.attr[self.first_keys[i]] = k + " break " + t
                                                    # print("successfully update", self.attr)
                                                    break
                                        except IndexError:
                                            pass
                            for k, v in DATE_PATTERN.items():
                                logging.debug("checking element date %s , %s, %s" % (k, v, element.get_text()))
                                potential_date = re.findall(v, element.get_text())
                                
                                if potential_date:
                                    self.potential_date_list.update(potential_date)
            logging.debug("this is dates %s "% self.potential_date_list)
            return self.attr

        def get_report_issue_date(self, pattern=None, guess=True):
            """

            :param pattern: dd/mm/yyyy, yyyy/mm/dd, mm/dd/yyyy
            :param guess: False when pattern is given, if True, try to guess the pattern
            :return: date type
            """

            def get_date_numbers(*args, date_tuple):
                """

                :param args: 0:mm int, 1:dd int, 2:yyyy int
                :param date_tuple: ('29', '.', '11', '.', '2018')
                :return: mm, dd, yyyy
                """
                _temp_list = [i.strip() for i in date_tuple if i not in string.punctuation]
                _temp_list = [i for i in _temp_list if i]
                if len(_temp_list) == 3: 
                    mm = int(_temp_list[args[0]])
                    dd = int(_temp_list[args[1]])
                    yyyy = int(_temp_list[args[2]])
                    try:
                        assert 0 < mm <= 12, "month should not greater than 12 or smaller than 1"
                        assert 0 < dd <= 31, "days should not smaller than 1 or greater than 31"
                        assert 2000 < yyyy < datetime.now().year +1, "years out of range"
                    except AssertionError:
                        # default value returned
                        return 1, 1, 2000   
                    return mm, dd, yyyy
                else:
                    return 1, 1, 2000

            def get_date_alpha(*args, date_tuple):
                """
                :param args: 0:str Oct, 1:dd int, 2: yyyy int
                :param date_tuple: ('Oct', '-', '15', '-', '2018')
                :return: mm, dd, yyyy
                """
                _temp_list = [i.strip() for i in date_tuple if i not in string.punctuation]
                _temp_list = [j for j in _temp_list if j]
                logging.debug("checking %s " % _temp_list)
                try:
                    # MONTH is a dictionary
                    mm = MONTH[_temp_list[args[0]]]
                except KeyError:
                    logging.debug("{%s} is not a valid month"%(_temp_list[args[0]]))
                    try:
                        int(_temp_list[args[0]])
                        # call last function
                        return get_date_numbers(args[1], args[0], args[2], date_tuple=date_tuple)
                    except ValueError:
                        return 0, 0, 0
                dd = int(_temp_list[args[1]])
                yyyy = int(_temp_list[args[2]])
                
                assert 0 < mm <= 12, "month should not greater than 12 or smaller than 1"
                assert 0 < dd <= 31, "days should not smaller than 1 or greater than 31"
                assert 2000 < yyyy < datetime.now().year +1, "years out of range"
                return mm, dd, yyyy

            def compare(date_tuple1, date_tuple2):
                _seq = [2, 0, 1]
                # date {mm, dd, yyyy}=> first compare year then month finally day
                while len(_seq) > 0:
                    _current_index = _seq.pop(0)
                    if date_tuple1[_current_index] < date_tuple2[_current_index]:
                        
                        return date_tuple2
                    elif date_tuple1[_current_index] > date_tuple2[_current_index]:
                        return date_tuple1
                    else:
                        logging.debug("same date piece pass")
                        pass
                # only come to this end if all same (all goes else block)
                return date_tuple1

            if self.potential_date_list:
                if pattern:
                    # pattern :str 
                    _temp_parttern_list = pattern.split('/')
                    alpha = False
                    try:
                        # get month idx
                        mm = _temp_parttern_list.index('mm')
                    except ValueError:
                        alpha = True
                        # oo repsents 'Jan','Feb'...
                        mm = _temp_parttern_list.index('oo')
                    dd = _temp_parttern_list.index('dd')
                    yyyy = _temp_parttern_list.index('yyyy')
                    issue_date = (0, 0, 0)

                    for i in list(self.potential_date_list):
                        if alpha:
                            # alpha indicate if month represents by letters
                            temp = get_date_alpha(mm, dd, yyyy, date_tuple=i)
                            issue_date = compare(issue_date, temp)
                        else:
                            temp = get_date_numbers(mm, dd, yyyy, date_tuple=i)
                            issue_date = compare(issue_date, temp)
                    try:
                        assert issue_date != (0, 0, 0), "not finding correct issue date, for {%s}" % self.potential_date_list
                    except AssertionError:
                        return "not found"
                    # datetime.strftime("%m/%d/%Y")
                    report_issue_date = datetime(month=issue_date[0], day=issue_date[1], year=issue_date[2])
                    self.attr['report_issue_date'] = report_issue_date.strftime("%m/%d/%Y")
                    return report_issue_date
                else:
                    # guess part,
                    pass
            else:
                return "not found"

        def refine_lab(self):
            pattern = r"\((.*?)\)"
            if len(self.attr['lab']) > 25:
                _current = self.attr['lab']
                try:
                    region = re.search(pattern, _current).group().strip("(|)")
                    lab = _current.partition(region)[0][:3]
                except AttributeError:
                    lab, region = _current.partition("break")[0][:3], _current.partition("break")[2][:10]
                self.attr['lab'] = (lab, region)
                return lab, region

        def refine(self, *args):
            try:
                if "break" in self.attr[args[0]]:
                    self.attr[args[0]] = keyword_modify(self.attr[args[0]].partition("break")[2].strip(),
                                                        special=True).strip()
            except KeyError:
                pass

        def refine_vender(self):
            # "break" is the separator mark
            try:
                if "break" in self.attr['factory_name']:
                    self.attr['factory_name'] = self.attr['factory_name'].partition("break")[2].strip()
                if "Previous Test" in self.attr['factory name']:
                    self.attr['factory_name'] = self.attr['factory_name'].partition("Previous Test")[0].strip()
            except KeyError:
                pass
            try:
                if "break" in self.attr['vendor_name']:
                    self.attr['vendor_name'] = self.attr['vendor_name'].partition("break")[2].strip()
                if "Test Type" in self.attr['vendor_name']:
                    self.attr['vendor_name'] = self.attr['vendor_name'].partition("Test Type")[0].strip()
            except KeyError:
                pass


class PDFExtractor:
    """
    put all piece together
    """

    def __init__(self, fp, target=None):
        self.PAGE_DICT = {}
        self.DICT_PAGE = {}
        self.fp = fp
        self.parser = PDFParser(fp)
        self.document = PDFDocument(self.parser)
        self.rsrcmgr = PDFResourceManager()
        self.laparams = LAParams(line_margin=LINE_MARGIN, char_margin=CHAR_MARGIN, all_texts=True, boxes_flow=0.5)
        self.pageagg = PDFPageAggregator(self.rsrcmgr, laparams=self.laparams)
        # self.textdevice = PDFTextDevice(self.pageagg)
        self.interpreter = PDFPageInterpreter(self.rsrcmgr, self.pageagg)
        if not self.document.is_extractable:
            raise PDFTextExtractionNotAllowed
        self.target_page = target
        self.begin_page = None
        self.general_info = None
        for pnumber, page in enumerate(PDFPage.get_pages(self.fp), start =1):
            self.PAGE_DICT[pnumber] = page.pageid
            self.DICT_PAGE[page.pageid] = pnumber
        logging.debug("Page Dict: %s" % self.PAGE_DICT)
        logging.debug("Page Dict reverse : %s" % self.DICT_PAGE)

    def collect_first_page_info(self):
        # changing and no longer used 
        temp_laparams = LAParams(line_margin=0.3, char_margin=1.5, all_texts=True, boxes_flow=0.5)
        temp_pageagg = PDFPageAggregator(self.rsrcmgr, laparams=temp_laparams)
        temp_interpreter = PDFPageInterpreter(self.rsrcmgr, temp_pageagg)
        self.general_info = {'report_number': None, 'protocol': None, 'vendor_name': None, 'factory_name': None, 'lab': None}
        for p in PDFPage.get_pages(self.fp):
            if p.pageid == self.PAGE_DICT[1]:
                temp_interpreter.process_page(p)
                layout = temp_pageagg.get_result()

                try:
                    l = FirstPageInfo(p, layout=layout, media_bbox=p.mediabox, attr=self.general_info)
                    l()
                except TypeError:
                    break
        return self.general_info

    def get_begin_page(self, check_list):
        """
        larger params makes it quicker
        :param check_list:  BEGIN KEY
        :return:
        """
        temp_laparams = LAParams(line_margin=0.3, char_margin=1.5, all_texts=True, boxes_flow=0.5)
        temp_pageagg = PDFPageAggregator(self.rsrcmgr, laparams=temp_laparams)
        temp_interpreter = PDFPageInterpreter(self.rsrcmgr, temp_pageagg)

        @timeout(5)
        def check_begin_keys(page, interpreter, pageagg, check_list):
            interpreter.process_page(page)
            layout = pageagg.get_result()
            found = False
            for element in layout:
                logging.debug("Currently checking : %s" % element)
                if isinstance(element, (LTTextBox, LTTextLine)):
                    for item in check_list:
                        if item in keyword_modify(element.get_text()):
                            found = True
                            logging.info('found the page: %s\n ' % page)
                            break
                        elif element.get_text() == SKIP_KEYS[0]:
                            # NOT Found but definately exceed the conclusion table begin page
                            break
            return found

        page_no = -1
        for p in PDFPage.get_pages(self.fp):
            if p.pageid != self.PAGE_DICT[1]:
                try:
                    m = check_begin_keys(p, interpreter=temp_interpreter, pageagg=temp_pageagg, check_list=check_list)
                    if m:
                        page_no = self.DICT_PAGE[p.pageid]
                        logging.info("Find begin page %d" % page_no)
                        break
                except TimeoutError:
                    logging.debug("Time out for detecting begin page")
                    break
        
        self.begin_page = page_no
        

    @timeout(5)
    def process_page(self):
        try:
            if self.target_page:
                for p in PDFPage.get_pages(self.fp):
                    for tp in self.target_page:
                        if p.pageid == self.PAGE_DICT[tp]:
                            self.interpreter.process_page(p)
                            layout = self.pageagg.get_result()
                            yield PageExtractor(layout=layout, pageid=p.pageid,page_number = self.DICT_PAGE[p.pageid], cache=True)
            else:
                start = self.begin_page            
                for p in PDFPage.get_pages(self.fp):
                    if self.DICT_PAGE[p.pageid] >= start:
                        self.interpreter.process_page(p)
                        layout = self.pageagg.get_result()
                        yield PageExtractor(layout=layout, pageid=p.pageid, page_number = self.DICT_PAGE[p.pageid],cache=True)
                    else:
                        pass

        except TimeoutError:
            logging.debug("Process_page time out")

    @timeout(5)
    def iter_page(self, start=None, end=None, single=None):
        """
        real page number
        start : int ,
        end : int,
        single : int,
        {start, end} : {(yes , no), (yes , yes), (no, yes), (no, no)}
        :param start: page to start to analysis , yield layouts if None, start from first page
        :param end: page to end , if None end till last page
        :param single: only yield single page, if None, iter from start to end
        :return: PageExtractor obj
        """
        if any((start, end)):
            assert isinstance(single, type(None))
            for p in PDFPage.get_pages(self.fp):
                if start and not end:
                    if self.DICT_PAGE[p.pageid] >= start:
                        self.interpreter.process_page(p)
                        layout = self.pageagg.get_result()
                        yield PageExtractor(layout=layout, pageid=p.pageid, page_number = self.DICT_PAGE[p.pageid],cache=True)
                elif start and end:
                    if start <= self.DICT_PAGE[p.pageid] <= end:
                        self.interpreter.process_page(p)
                        layout = self.pageagg.get_result()
                        yield PageExtractor(layout=layout, pageid=p.pageid,page_number = self.DICT_PAGE[p.pageid], cache=True)
                elif not start and end:
                    if 1 <= self.DICT_PAGE[p.pageid] <= end:
                        self.interpreter.process_page(p)
                        layout = self.pageagg.get_result()
                        yield PageExtractor(layout=layout, pageid=p.pageid,page_number = self.DICT_PAGE[p.pageid],cache=True)
        else:
            assert isinstance(single, int), "if not start , end , single must be given"
            for p in PDFPage.get_pages(self.fp):
                if self.DICT_PAGE[p.pageid] == single:
                    self.interpreter.process_page(p)
                    layout = self.pageagg.get_result()
                    yield PageExtractor(layout=layout, pageid=p.pageid,page_number = self.DICT_PAGE[p.pageid], cache=True)


class PDFExtractorFitz:

    def __init__(self, fname):
        self.doc = fitz.open(fname)

    def __iter__(self):
        for i in range(len(self.doc)-100):
            yield PageExtractorFitz( page_doc = self.doc[i]) 

    def get_page(self, number):
        return PageExtractorFitz(page_doc = self.doc[number])


class PageExtractorFitz:
    def __init__(self, page_doc):
        """
        page#: int 0 base
        page: fitz.Page
         number => page.number 
         rect => page.rect | MediaBox  left top(0,0) right bottom(max, max)

        """
        self.page = page_doc
        self._header = None
        self._body = None
        self._footer = None

    @property
    def header(self):
        return self._header

    @header.setter
    def header(self, rate = 0.2):
        # header is top 20% of the page
        x0,y0,x1,y1 =  self.page.rect.x0, self.page.rect.y0, self.page.rect.x1, self.page.rect.y1 *rate
        self._header = fitz.Rect(x0,y0,x1,y1).round()

    @property
    def body(self):
        return self._body

    @body.setter
    def body(self, top = 0.2, bottom = 0.2):
        x0,y0,x1,y1 =  self.page.rect.x0, self.page.rect.y1 * top, self.page.rect.x1, self.page.rect.y1 * (1-bottom)
        self._body = fitz.Rect(x0,y0,x1,y1).round()
    
    @property
    def footer(self):
        return self._footer

    @footer.setter
    def footer(self, rate = 0.2):
        x0,y0,x1,y1 =  self.page.rect.x0, self.page.rect.y0 * (1-rate), self.page.rect.x1, self.page.rect.y1
        self._footer = fitz.Rect(x0,y0,x1,y1).round()
            
    def search_for(self, text, words):
        rect_list = []
        for w in words:
          
            if text.lower() in w[4].lower():
                rect_list.append(w)
              
        return rect_list

    def force_search(self, text, words):
        pass
        
    def check_blocks(self, words, rect):
        """
        words => page.getText("block")
        rect => Rect
        """
        # build my sublist of words contained in given rectangle
        mywords = [w for w in words if fitz.Rect(w[:4]) in rect]

        return mywords

    def recover(self, words, rect):
        """ Word recovery.

        Notes:
            Method 'getTextWords()' does not try to recover words, if their single
            letters do not appear in correct lexical order. This function steps in
            here and creates a new list of recovered words.
        Args:
            words: list of words as created by 'getTextWords()'
            rect: rectangle to consider (usually the full page), check report number top 20% of the page
        Returns:
            List of recovered words. Same format as 'getTextWords', but left out
            block, line and word number - a list of items of the following format:
            [x0, y0, x1, y1, "word"]
        """
        # build my sublist of words contained in given rectangle
        mywords = [w for w in words if fitz.Rect(w[:4]) in rect]

        # sort the words by lower line, then by word start coordinate
        mywords.sort(key=itemgetter(3, 0))  # sort by y1, x0 of word rectangle

        # build word groups on same line
        grouped_lines = groupby(mywords, key=itemgetter(3))

        words_out = []  # we will return this

        # iterate through the grouped lines
        # for each line coordinate ("_"), the list of words is given
        for _, words_in_line in grouped_lines:
            for i, w in enumerate(words_in_line):
                if i == 0:  # store first word
                    x0, y0, x1, y1, word = w[:5]
                    continue

                r = fitz.Rect(w[:4])  # word rect

                # Compute word distance threshold as 20% of width of 1 letter.
                # So we should be safe joining text pieces into one word if they
                # have a distance shorter than that.
                threshold = r.width / len(w[4]) / 5
                if r.x0 <= x1 + threshold:  # join with previous word
                    word += w[4]  # add string
                    x1 = r.x1  # new end-of-word coordinate
                    y0 = max(y0, r.y0)  # extend word rect upper bound
                    continue

                # now have a new word, output previous one
                words_out.append([x0, y0, x1, y1, word])

                # store the new word
                x0, y0, x1, y1, word = w[:5]

            # output word waiting for completion
            words_out.append([x0, y0, x1, y1, word])

        return words_out

class Searcher(BasicDict):
    def __iter__(self, idx, key):
        super().__init__(idx = idx, key= key)

    def page_search_for(self, page, flag = 1):
        """
        flag == 1, block search
        """
        wordsout = []
        if flag == 1:
            words = page.page.getTextWords()
            print("words are here", words)
        if len(self.keywords) >0:
            # keywords exits
            if self.bbox:
                # bbox exits    
                search_area = fitz.Rect(self.bbox[:])
                words = page.check_blocks(words = words, rect = search_area)

            for text in self.keywords:  # [ str, str]
                wordsout += page.search_for(text = text, words = words)
        if len(wordsout) > 0:
            if self.force:
                for g in wordsout:
                    if any([p.lower() in g[4].lower() for p in self.force]):            
                        return set(wordsout)
        if self.force:
            for text in self.force:
                print(text)
                wordsout += page.search_for(text = text, words = words)
        else:
            logging.warning("Sorry no force search items & could not find matched items for {%s}" % self.item)
        return set(wordsout) 
        

if __name__ == '__main__':
    start_timestamp = time.time()
    """path = PATH['path']
    sub_path = PATH['sub_path'][5]

    doc = Document(path=path, sub_path=sub_path, doc_id=8)
    doc()
    
    tag_name = doc.get_file_name()

    # os.startfile(doc.file_path)
    pages = tag_name.split('-')
    # target_pages = [9]
    fp = open(doc.file_path, 'rb')
 
    pdf_ss = PDFExtractor(fp)
    genenral_info = pdf_ss.collect_first_page_info()
    pdf_ss.get_begin_page(BEGIN_KEYS)
    st = pdf_ss.process_page()

    p = genenral_info['protocol']
    guessed = get_protocol(p, path=PROTOCOLS)
   
    select_protocol = user_select_from_list(name='protocol', item=guessed)

    protocol_df = pd.read_excel(PROTOCOLS + select_protocol)

    cat = Mapper(df=protocol_df)
    cat.predefine_checker_list()
    wb = Workbook()
    g = wb.active
    g.title = 'general info'
    g.append(list(genenral_info.keys()))
    g.append([str(i) for i in list(genenral_info.values())])
    wb.create_sheet(title="result")
    count = 0

    while count < 9:
        apage = next(st)
        apage()
        ws = wb.create_sheet()
        apage.page_to_excel(ws=ws, ref=pdf_ss.DICT_PAGE)
       
        body = apage.get_body()
        mouse = predefine_(body)
        cat.verify_item(page_number=pdf_ss.DICT_PAGE[apage.pageid], pdf_df=mouse)
        count += 1

    wb.save(os.getcwd() + '/excels/' + tag_name + 'newly' + '.xlsx')
    append_df_to_excel(os.getcwd() + '/excels/' + tag_name + 'newly' + '.xlsx', cat.df, sheet_name="result")
    end_time = time.time()"""
    fname = os.getcwd()
 

