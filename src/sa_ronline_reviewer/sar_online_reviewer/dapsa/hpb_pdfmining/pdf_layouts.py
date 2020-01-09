import re
import copy
import numpy as np
import pandas as pd
from pdfminer.pdfpage import PDFPage
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from .excel_styles import get_style, STYLES, style_range, adjust_range_size
import logging

PTS = 0.2
try:
    highlight = get_style("highlight", style_dict=STYLES)
    normal = get_style("normal", style_dict=STYLES)
except ValueError:
    pass

cell_alighment = Alignment(horizontal="center", vertical="center")


def approxiamtion(xc, x0, width=1.5, pts=PTS):
    return abs(xc - x0) < width * (1 + pts)

def close_enough(a,b):
    dist = np.sqrt(np.square(a[0]-b[0])+np.square(a[1]-b[1]))
    return dist <= (1+PTS) * 1.5

def inside(bigger_obj, smaller_obj, width=3.0, pts=0.2):
    matrix = np.array(bigger_obj.bbox) >= np.array(smaller_obj.bbox)
    distinct = np.array(bigger_obj.bbox) >= (np.array(smaller_obj.bbox) +
                                             np.array((width * (1 + pts), width * (1 + pts), 0 - width * (1 + pts),
                                                       0 - width * (1 + pts))))
    if not any(matrix[0:2]) and all(matrix[2:]):
        return True
    else:
        if not any(distinct[0:2]) and all(distinct[2:]):
            return True
        else:
            return False


def find_next_neighbor(media_bbox, _objs, obj, width=3, pts=0.2, direction=1):
    """a generator, beware for direction 2, and 4 the
    obj will return from furthest to cloestest"""

    class Area:
        """bbox := tuple(float,float,float,float)"""

        def __init__(self):
            self.bbox = None

        def assign_bbox(self, bbox):
            self.bbox = bbox

    current = Area()
    if direction == 1:
        'go right'
        target_bbox = (obj.bbox[2], obj.bbox[1], media_bbox[2], obj.bbox[3])
        current.assign_bbox(target_bbox)
    elif direction == 2:
        'go left, it will return from left to right, sort outside'
        target_bbox = (media_bbox[0], obj.bbox[1], obj.bbox[0], obj.bbox[3])
        current.assign_bbox(target_bbox)
    elif direction == 3:
        'go bottom'
        target_bbox = (obj.bbox[0], media_bbox[1], obj.bbox[2], obj.bbox[1])
        current.assign_bbox(target_bbox)
    elif direction == 4:
        'go top'
        target_bbox = (obj.bbox[0], obj.bbox[3], obj.bbox[2], media_bbox[3])
        current.assign_bbox(target_bbox)
    for element in _objs:
        if isinstance(element, type(obj)):
            if element.bbox[2] - element.bbox[0] > 3:
                'it is still a good protection, for most cells has a width greater than 3'
                if inside(bigger_obj=current, smaller_obj=element, width=width, pts=pts):
                    yield element


class Point:
    """
    a child Point represent the middle point of
    two points from a LTCurve, it is the deep
    tag indicate that if two lines belong to a
    table
    """
    pagePoints = {}
    pageMiddlePoints = {}

    def __init__(self, x, y, pageid, point_id, line_width=2, is_child=False):
        self.x = x
        self.y = y
        self.pageid = pageid
        self.point_id = point_id
        self.line_width = line_width
        self.is_child = is_child
        self.parents = []
        self.children = []

    def get_point_id(self):
        return self.point_id

    def assign_child(self, child):
        """
         id   location
         1     up
         2     right
         3     down
         4     left
         5     right up
         6     right down
         7     left down
         8     left up
        """

        def define_direction(x, y, child, linewidth):
            if approxiamtion(x, child.x) and not approxiamtion(y, child.y):
                if child.y - y > linewidth:
                    return 1
                elif child.y - y < -linewidth:
                    return 3
            elif not approxiamtion(x, child.x) and not approxiamtion(y, child.y):
                if child.x - x > linewidth:
                    if child.y - y > linewidth:
                        return 5
                    elif child.y - y < -linewidth:
                        return 6
                elif child.x - x < -linewidth:
                    if child.y - y < -linewidth:
                        return 7
                    elif child.y - y > linewidth:
                        return 8
            elif not approxiamtion(x, child.x) and approxiamtion(y, child.y):
                if child.x - x > linewidth:
                    return 2
                elif child.x - x < -linewidth:
                    return 4
            else:
                return 1

        assert child.is_child

        if self.is_same_point(child):
            child.parents.append(-999)
            return
        else:
            pipid = define_direction(self.x, self.y, child, linewidth=self.line_width * (1 + PTS))
            child.parents.append(self.point_id)
            self.children.append((pipid, child))
        return

    def is_same_point(self, other, pts=0.2):
        dist = self.get_dist(other)
        if dist <= (1 + pts) * self.line_width:
            return True
        else:
            return False

    def get_dist(self, other):
        dist = np.sqrt(np.square(abs(self.x - other.x)) + np.square(abs(self.y - other.y)))
        return dist

    def is_starter(self):
        pass

    def __repr__(self):
        return ('<pageid: %s , id: %s, coordinates: (%s %s), childpoint: %s , parents: %s >' %
                (self.pageid, self.point_id, self.x, self.y, self.is_child, self.parents))

    @classmethod
    def set_page_points(cls, pageid):
        cls.pagePoints[pageid] = []
        cls.pageMiddlePoints[pageid] = []

    def add_to_page_points(self, pageid):
        if not self.is_child:
            Point.pagePoints[pageid].append(self)
        else:
            if not all(i == -999 for i in self.parents):
                Point.pageMiddlePoints[pageid].append(self)

    @classmethod
    def clean_points(cls, pageid, textboxlist):
        # it wont cause trouble to closed tables
        def check_points_valiadation(point_list, textboxlist):
            # remove all points in a textbox e.g a html address with bottom line
            valid_points = []
            un_valid_points = []
            for point in point_list:
                un_valid = False
                for t in textboxlist:
                    if all([point.x >= t.bbox[0], point.x <= t.bbox[2],
                            point.y >= t.bbox[1], point.y <= t.bbox[3]]):
                        un_valid = True
                if un_valid:
                    un_valid_points.append(point)
                else:
                    valid_points.append(point)
            return valid_points

        cls.pagePoints[pageid] = check_points_valiadation(cls.pagePoints[pageid], textboxlist=textboxlist)

    @classmethod
    def purge(cls):
        """
        purge all
        :return:
        """
        cls.pagePoints = {}
        cls.pageMiddlePoints = {}


class Cluster:
    """
    a cluster is a collection of points
    all points in same cluster should
    be defined as 'same' to one another
    and a cluster should contain all
    same points in a page and given a
    coordinates represent the cluster
    it should be a graph data type and
    the represent value should be the
    centre of the cluster
    """
    pageCluster = {}
    pageSortedClusters = {}
    pageRoots = {}
    pageCells = {}
    starters = {}

    def __init__(self):
        self.key = (-1, -1)
        self.value = set()
        self.pageid = None
        self.cluster_id = None
        self.point_id = set()
        self.mid_targets = []
        self.target_info = []
        self.s1_matrix = []
        self.root_id = None
        self.tableRoot = False
        self.rowChild = None
        self.colChild = None
        self.rowParent = None
        self.colParent = None
        self.color = 'white'
        self.byte_mark = 0

    @classmethod
    def purge(cls):
        cls.pageCluster = {}
        cls.pageSortedClusters = {}
        cls.pageRoots = {}
        cls.pageCells = {}
        cls.starters = {}

    def same_cluster(self, other):
        # compare two clusters
        if approxiamtion(self.key[0], other.key[0]) \
                and approxiamtion(self.key[1], other.key[1]):
            return True
        else:
            return False

    def get_key(self):
        return self.key

    def get_value(self):
        return self.value

        # need refine

    def is_acceptable(self, other):
        acceptable = False
        check = [other.is_same_point(i) for i in self.value]
        if all(check):
            acceptable = True
        return acceptable

    def __repr__(self):
        return ('<pageid: %s , cluster_id: %s, '
                'cluster_key: %s, number_of_points: %s ,'
                'pointsID: %s ,'
                'number_of_children = %s>\n' %
                (self.pageid, self.cluster_id, self.key,
                 len(self.value), self.point_id,
                 len([item.children for item in self.value if item.children])))

    def fill_(self, point):
        """
        :param point:
        :return:
        """

        class Fakepoint:
            def __init__(self, x, y):
                self.x = x
                self.y = y

        if len(self.value) == 0:
            self.key = (point.x, point.y)
            self.value.update([point])
            self.pageid = point.pageid
            self.point_id.update([point.point_id])
        else:
            self.value.update([point])

            self.calculate_key()
            temp_fake_point = Fakepoint(x=self.key[0], y=self.key[1])
            if all(item.is_same_point(temp_fake_point) for item in self.value):
                # keep the new point, update points id
                self.get_points_id()
            else:
                # remove new added point
                self.value.remove(point)
                self.calculate_key()

    def calculate_key(self):
        # calculate new key value for each time a new point added in
        self.key = (sum(i.x for i in self.value) / len(self.value),
                    sum(i.y for i in self.value) / len(self.value))

    def get_points_id(self):
        for i in self.value:
            self.point_id.update([i.point_id])
        return self.point_id

    def is_row_child(self):
        if self.rowParent:
            return self.rowParent.rowChild == self

    def is_col_child(self):
        if self.colParent:
            return self.colParent.colChild == self

    def is_root(self):
        if self.s1_matrix[0] == [0, 1, 1, 0]:
            self.tableRoot = True
        return self.tableRoot

    def set_s1_matrix(self):
        v = []
        for i in range(8):
            v.append((0, [(-1, 0)]))
        logging.info("setting s1 matrix for %s as %s", self.cluster_id, v)
        for m in self.target_info:
            logging.info(" %s m: %s in %s 's target_info", self.set_s1_matrix.__name__, m, self.cluster_id)
            # m = (direction, [(cluster id, 0)]
            v[m[0] - 1] = m
        self.s1_matrix = [[int(v[i][0] != 0) for i in range(len(v))], [i[1][0] for i in v]]
        # 这里需要把list的前4个转为二进制代码 ， 并将相对应的起始态存入dict
        print(self.cluster_id, self.s1_matrix)

    # need refine
    def get_middle(self):
        # connect with other clusters
        for item in self.value:
            logging.info(" %s item is %s in %s", self.get_middle.__name__, item, self.value)
            if item.children:
                self.mid_targets += item.children
                for t in item.children:
                    self.target_info.append((t[0], [n for n in t[1].parents if n not in self.point_id]))
                logging.info("traget info: %s", self.target_info)
            else:
                logging.info(" %s checking  %s no children, ", self.get_middle.__name__, item)

        # O(n2) max 4*N N is all points in one page
        for s in self.target_info:
            found = False
            for j in self.pageSortedClusters[self.pageid]:
                if j != self:
                    for m in range(len(s[1])):
                        if s[1][m] in j.point_id:
                            s[1][m] = (j.cluster_id, m)
                            logging.info("s is %s, current points %s", s, self.point_id)
                            found = True
            if not found:
                logging.warning("_________________________________________")
                logging.warning("not found matched cluster for %s,", s)
                logging.warning("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.set_s1_matrix()
        return

    def net_up(self):
        if self.s1_matrix[0][0] == 1:
            self.colParent = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][0][0]]
        if self.s1_matrix[0][1] == 1:
            self.rowChild = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][1][0]]
        if self.s1_matrix[0][2] == 1:
            self.colChild = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][2][0]]
        if self.s1_matrix[0][3] == 1:
            self.rowParent = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][3][0]]

    def do_detect(self):
        return sum(self.s1_matrix[0][1:3])

    def do1(self):
        # row child judgement put outside this function
        k = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][1][0]].cluster_id
        m = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][1][0]].s1_matrix[1][0][0]
        if m != -1:
            return k
        else:
            return Cluster.pageSortedClusters[self.pageid][k].do1()

    def do2(self, key):
        # row child judgement put outside this function
        k = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][1][0]].cluster_id
        m = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][1][0]].s1_matrix[1][2][0]
        
        if Cluster.pageSortedClusters[self.pageid][k].do_detect() == 2:
            if Cluster.pageSortedClusters[self.pageid][k].color == 'white':
                logging.debug("do 2 adding {%s} color is {%s} " % (k, Cluster.pageSortedClusters[self.pageid][k].color))
                Cluster.starters[key].update([k])
        if m != -1:
            return k
        else:
            try:
                assert Cluster.pageSortedClusters[self.pageid][k].s1_matrix[1][1][0] != -1, "cannot do 2, %s" % self
                logging.debug("detect s1 matrix [%s]" %sum(Cluster.pageSortedClusters[self.pageid][k].s1_matrix[0]))
                return Cluster.pageSortedClusters[self.pageid][k].do2(key)
            except AssertionError:
                logging.debug("return %s " % Cluster.pageSortedClusters[self.pageid][k].s1_matrix[1][3][0])
                return -1

    def do3(self, key):
        # row child judgement put outside this function
        try:
            k = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][2][0]].cluster_id
            m = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][2][0]].s1_matrix[1][1][0]
            logging.debug("Do2 : k : {%s}{%s}" %(k, Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][1][0]].s1_matrix))
            if Cluster.pageSortedClusters[self.pageid][k].do_detect() == 2:
                if Cluster.pageSortedClusters[self.pageid][k].color == 'white':
                    logging.debug("do 3 adding {%s} color is {%s}" % (k, Cluster.pageSortedClusters[self.pageid][k].color))
                    Cluster.starters[key].update([k])
            if m != -1:
                return k
            else:
                try:
                    assert Cluster.pageSortedClusters[self.pageid][k].s1_matrix[1][2][0] != -1, "cannot do 3, %s" % self
                    return Cluster.pageSortedClusters[self.pageid][k].do3(key)
                except AssertionError:
                    logging.debug("return %s " % Cluster.pageSortedClusters[self.pageid][k].s1_matrix[1][1][0])
                    return -1
        except TypeError:
            logging.warning(" %s , get error on matrix %s ", self, self.s1_matrix)

    def do4(self):
        # row child judgement put outside this function
        k = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][2][0]].cluster_id
        m = Cluster.pageSortedClusters[self.pageid][self.s1_matrix[1][2][0]].s1_matrix[1][3][0]
        if m != -1:
            return k
        else:
            return Cluster.pageSortedClusters[self.pageid][k].do4()

    @classmethod
    def page_cluster_is_empty(cls, pageid):
        return pageid not in cls.pageCluster.keys()

    @classmethod
    def set_page_cluster(cls, pageid):
        if cls.page_cluster_is_empty(pageid):
            cls.pageCluster[pageid] = set()
            cls.pageRoots[pageid] = set()
            cls.pageCells[pageid] = {}

    # starting point
    @classmethod
    def group_cluster(cls, pageid, cleaned_points):
        for point in cleaned_points:
            if point.pageid == pageid:
                if cls.page_cluster_is_empty(pageid):
                    cls.set_page_cluster(pageid)
                if cls.pageCluster[pageid]:
                    found = False
                    for cluster in cls.pageCluster[pageid]:
                        cluster.fill_(point)
                        if point.point_id in cluster.point_id:
                            found = True
                            break
                    if not found:
                        new_cluster = Cluster()
                        new_cluster.fill_(point)
                        cls.add_page_cluster(new_cluster)
                else:
                    new_cluster = Cluster()
                    new_cluster.fill_(point)
                    cls.add_page_cluster(new_cluster)
        return

    @classmethod
    def add_page_cluster(cls, cluster):
        if not cls.page_cluster_is_empty(cluster.pageid):
            cls.pageCluster[cluster.pageid].update([cluster])
        else:
            cls.pageCluster[cluster.pageid] = set()
            cls.pageCluster[cluster.pageid].update([cluster])

    @classmethod
    def assign_id(cls, pageid):
        # O(page*n), better use a page parameter, or assign id after caching?
        for j, v in enumerate(list(cls.pageSortedClusters[pageid])):
            v.cluster_id = j
            logging.info(v)
        return

    @classmethod
    def sort_page_cluster(cls, pageid):
        cls.pageSortedClusters[pageid] = sorted(sorted(list(cls.pageCluster[pageid]), key=lambda x: x.key[0]),
                                                key=lambda x: round(x.key[1]), reverse=True)

    @classmethod
    def locate_roots(cls, pageid):
        # use after sort clusters
        for c in cls.pageSortedClusters[pageid]:
            if c.s1_matrix[0][0:4] == [0, 1, 1, 0]:
                cls.pageRoots[pageid].update([c.cluster_id])

    @classmethod
    def do_group(cls, pageid):
        for c in cls.pageSortedClusters[pageid]:
            c.net_up()
        return

    @classmethod
    def remove_header(cls):
        pass

    @classmethod
    def do_cell(cls, pageid):
        # f = det + do2+do3+do_compare start from root
        cls.starters = {}
        while len(cls.pageRoots[pageid]) > 0:
            c_root = cls.pageRoots[pageid].pop()
            cls.pageCells[pageid][c_root] = []
            # all starters comes form the c_roots
            cls.starters[c_root] = set()
            logging.info("current starter is %s " %cls.starters)
            cls.starters[c_root].update([c_root])
            logging.info("this is root %s" % c_root)
            while len(cls.starters[c_root]) > 0:
                c_starter = cls.starters[c_root].pop()
                if cls.pageSortedClusters[pageid][c_starter].do_detect() == 2:
                    m = cls.pageSortedClusters[pageid][c_starter].do2(c_root)
                    n = cls.pageSortedClusters[pageid][c_starter].do3(c_root)
                    mbbox = cls.pageSortedClusters[pageid][m].key
                    nbbox = cls.pageSortedClusters[pageid][n].key
                    assumption = (mbbox[0], nbbox[1])
                    if not any([m == -1, n == -1]):
                        k = cls.pageSortedClusters[pageid][m].do4()
                        t = cls.pageSortedClusters[pageid][n].do1()
                        if k == t:
                            s = cls.pageSortedClusters[pageid][k].s1_matrix[0][0:4]
                            x = {'pageid': pageid, 'table_root': c_starter == c_root, 'table_end': s == [1, 0, 0, 1],
                                 'root': c_starter, 'rtc': m, 'lbc': n, 'rbc': k}
                            # set to a cell
                            logging.info("adding cell starter:%s, m:%s, n:%s, t:%s"%(c_starter, m, n, t))
                            cls.pageCells[pageid][c_root].append(
                                Cell(x, is_root=x['table_root'], is_end=x['table_end']))
                            cls.pageSortedClusters[pageid][c_starter].color = 'yellow'
                        else:
                            kbbox = cls.pageSortedClusters[pageid][k].key
                            tbbox = cls.pageSortedClusters[pageid][t].key
                            if close_enough(kbbox,assumption):
                                s = cls.pageSortedClusters[pageid][k].s1_matrix[0][0:4]
                                x = {'pageid': pageid, 'table_root': c_starter == c_root,
                                     'table_end': s == [1, 0, 0, 1],
                                     'root': c_starter, 'rtc': m, 'lbc': n, 'rbc': k}
                                # set to a cell
                                logging.info("adding cell starter:%s, m:%s, n:%s, k:%s"%(c_starter, m, n, k))
                                cls.pageCells[pageid][c_root].append(
                                    Cell(x, is_root=x['table_root'], is_end=x['table_end']))
                                cls.pageSortedClusters[pageid][c_starter].color = 'yellow'
                            else:
                                if close_enough(tbbox,assumption):
                                    s = cls.pageSortedClusters[pageid][t].s1_matrix[0][0:4]
                                    x = {'pageid': pageid, 'table_root': c_starter == c_root,
                                         'table_end': s == [1, 0, 0, 1],
                                         'root': c_starter, 'rtc': m, 'lbc': n, 'rbc': t}
                                    # set to a cell
                                    logging.info("adding cell, %s, %s, %s, %s " % (c_starter, m, n, t))
                                    cls.pageCells[pageid][c_root].append(
                                        Cell(x, is_root=x['table_root'], is_end=x['table_end']))
                                    cls.pageSortedClusters[pageid][c_starter].color = 'yellow'
                                else:
                                    logging.warning("Incorrect cell id {table root : %s }, "
                                                    "{root : %s } , { (rtc: %s),(lbc: %s), (rbc: %s, %s)}",
                                                    c_starter == c_root,
                                                    c_starter, m, n, k, t)
                                    for c in cls.pageSortedClusters[pageid]:
                                        if close_enough(c.key,assumption):
                                            s = cls.pageSortedClusters[pageid][c.cluster_id].s1_matrix[0][0:4]
                                            x = {'pageid': pageid, 'table_root': c_starter == c_root,
                                                 'table_end': s == [1, 0, 0, 1],
                                                 'root': c_starter, 'rtc': m, 'lbc': n, 'rbc': t}
                                            # set to a cell
                                            logging.info("adding cell starter: %s , m: %s, n:%s, t:%s "%(c_starter, m, n, t))
                                            cls.pageCells[pageid][c_root].append(
                                                Cell(x, is_root=x['table_root'], is_end=x['table_end']))
                                            cls.pageSortedClusters[pageid][c_starter].color = 'yellow'


class Cell:
    """{pageid : {table_root:[cells]}}"""
    starters = {}
    pageCells = Cluster.pageCells

    def __init__(self, matrix, is_root=False, is_end=False):
        self.matrix = matrix
        self.pageid = matrix['pageid']
        self.cell_id = None
        self._height = -1
        self._width = -1
        self._upper_boundary = -1
        self._lower_boundary = -1
        self._left_boundary = -1
        self._right_boundary = -1
        self._centre = None
        self._bbox = None
        self._text = []
        self.is_root = is_root
        self.is_end = is_end
        self._row_color = ''
        self._col_color = ''
        self._row_children = set()
        self._col_children = set()
        self._row_parent = set()
        self._col_parent = set()
        self.virtual_matrix = None
        self.table_no = None
        # self._merge_relation

    def __call__(self, *args, **kwargs):
        self.upper_boundary = self.matrix
        self.lower_boundary = self.matrix
        self.left_boundary = self.matrix
        self.right_boundary = self.matrix
        self.height = (self._upper_boundary - self._lower_boundary)
        self.width = (self._right_boundary - self._left_boundary)
        self.centre = [self._left_boundary, self._right_boundary, self._lower_boundary, self._upper_boundary]
        self.bbox = (self._left_boundary, self._lower_boundary, self._right_boundary, self._upper_boundary)
        self.row_color = 'white'
        self.col_color = 'white'

    def virtual_matrix_builder(self, fulfilled_matrix):
        if self.virtual_matrix is not None:
            self.virtual_matrix += fulfilled_matrix
        else:
            self.virtual_matrix = fulfilled_matrix
        return

    @classmethod
    def purge(cls):
        cls.starters = {}
        cls.pageCells = Cluster.pageCells

    def check_virtual_matrix(self):
        # pending check if the matrix not contains [1, 0, 1, 1]
        return (self.virtual_matrix < 2).all()

    @property
    def row_color(self):
        return self._row_color

    @row_color.setter
    def row_color(self, new_color):
        self._row_color = new_color

    @property
    def col_color(self):
        return self._col_color

    @col_color.setter
    def col_color(self, new_color):
        self._col_color = new_color

    @property
    def row_parent(self):
        return self._row_parent

    @row_parent.setter
    def row_parent(self, obj):
        self._row_parent = obj

    @property
    def col_parent(self):
        return self._col_parent

    @col_parent.setter
    def col_parent(self, obj):
        self._col_parent = obj

    @property
    def row_children(self):
        return self._row_children

    @row_children.setter
    def row_children(self, media_bbox):
        k = find_next_neighbor(media_bbox=media_bbox, _objs=Cell.pageCells[self.pageid], obj=self, direction=1)
        while True:
            try:
                target = next(k)
                if approxiamtion(self.bbox[0], target.bbox[0], width=2, pts=0.2):
                    '3 is too wide'
                    self.row_children.update([target])
                    if self.row_color == 'white':
                        self.__setattr__('row_color', 'yellow')
                    else:
                        self.__setattr__('row_color', 'black')
                    target.row_parent = self
                    break
                else:
                    pass
            except StopIteration:
                print('not found next neighbor')
                break

    @property
    def col_children(self):
        return self._col_parent

    @col_children.setter
    def col_children(self, media_bbox):
        k = find_next_neighbor(media_bbox=media_bbox, _objs=Cell.pageCells[self.pageid], obj=self, direction=3)
        while True:
            try:
                target = next(k)
                if approxiamtion(self.bbox[0], target.bbox[0], width=2, pts=0.2):
                    '3 is too wide'
                    self.col_children.update([target])
                    if self.col_color == 'white':
                        self.__setattr__('col_color', 'yellow')
                    else:
                        self.__setattr__('col_color', 'black')
                    target.col_parent = self
                    break
                else:
                    pass
            except StopIteration:
                print('not found next neighbor')
                break

    def set_row_neightbor(self, obj):
        """
        TBD
        the property setting can only work when there are no merge cells, if a cell color remain white\
        after property setting, this function will set parents and children for it"""
        _root = self.matrix['root']
        _lbc = self.matrix['lbc']
        _rtc = self.matrix['rtc']
        _rbc = self.matrix['rbc']
        _root = self.matrix['root']
        _lbc = self.matrix['lbc']
        _rtc = self.matrix['rtc']
        _rbc = self.matrix['rbc']
        if self.is_root:
            'no parent then'
            pass
        elif self.is_end:
            'no children then'
            pass
        else:
            pass

    @property
    def upper_boundary(self):
        return self._upper_boundary

    @upper_boundary.setter
    def upper_boundary(self, matrix):
        self._upper_boundary = round(max(Cluster.pageSortedClusters[self.pageid][matrix['root']].key[1],
                                         Cluster.pageSortedClusters[self.pageid][matrix['rtc']].key[1]), 2)

    @property
    def lower_boundary(self):
        return self._lower_boundary

    @lower_boundary.setter
    def lower_boundary(self, matrix):
        self._lower_boundary = round(min(Cluster.pageSortedClusters[self.pageid][matrix['lbc']].key[1],
                                         Cluster.pageSortedClusters[self.pageid][matrix['rbc']].key[1]), 2)

    @property
    def left_boundary(self):
        return self._left_boundary

    @left_boundary.setter
    def left_boundary(self, matrix):
        self._left_boundary = round(min(Cluster.pageSortedClusters[self.pageid][matrix['lbc']].key[0],
                                        Cluster.pageSortedClusters[self.pageid][matrix['root']].key[0]), 2)

    @property
    def right_boundary(self):
        return self._right_boundary

    @right_boundary.setter
    def right_boundary(self, matrix):
        self._right_boundary = round(max(Cluster.pageSortedClusters[self.pageid][matrix['rtc']].key[0],
                                         Cluster.pageSortedClusters[self.pageid][matrix['rbc']].key[0]), 2)

    @property
    def height(self):
        return self._height

    @height.setter
    def height(self, h):
        self._height = round(h, 2)

    @property
    def width(self):
        return self._width

    @width.setter
    def width(self, w):
        self._width = round(w, 2)

    @property
    def centre(self):
        return self._centre

    @centre.setter
    def centre(self, value):
        self._centre = (round((value[1] - value[0]) / 2 + value[0], 2), round((value[3] - value[2]) / 2 + value[2], 2))

    @property
    def bbox(self):
        return self._bbox

    @bbox.setter
    def bbox(self, t):
        self._bbox = t

    @property
    def text(self):
        return self._text

    @text.setter
    def text(self, t):
        self._text += [t]

    def __repr__(self):
        return ('<pageid: %s ,'
                'table_no: %s '
                'cell_id: %s, '
                'matrix: %s ,'
                'bbox: %s ,'
                'centre: %s ,'
                'content: %s, '
                'vir_martix: %s ,'
                '>\n'
                % (self.pageid, self.table_no, self.cell_id,
                   self.matrix, self._bbox, self._centre, self.text,
                   self.virtual_matrix))

    @classmethod
    def assign_text(cls, pageid, textlist):
        temp = copy.deepcopy(textlist)
        while len(temp) > 0:
            current = temp.pop(0)
            for v in cls.pageCells[pageid].values():
                for c in v:
                    if inside(c, current):
                        c.text = current.get_text()


    @classmethod
    def check_text(cls, pageid):
        for v in cls.pageCells[pageid].values():
            for c in v:
                logging.info("the current cell {%s} contains text {%s}, clusters : {%s}", c.bbox, c.text, c.matrix)
                pass


class Table:
    """this will be a pandas dataframe with unique table id , maybe by page id and other parameters
    potentially  a table should provide {
    table name , table field ( column names) , row index and so on
    another issue here is for some row and columns with for instance visually same but not share precise cooridinates
    it can be reorganised here
    and last a table better provide a write into function to write a table to a excel workbook with sheet name
    equals table names (maybe the header or few words of left up text box)"""

    def __init__(self, bbox, pageid, doc, table_no, base_matrix, horizontal_list, vertical_list):
        self.bbox = bbox
        self.table_no = table_no
        self.pageid = pageid
        self.doc = doc
        self.base_matrix = base_matrix
        self.horizontal_list = horizontal_list
        self.vertical_list = vertical_list

    def __repr__(self):
        return '< pageid: %s, ' \
               'table_no: %s  ' \
               'bbox: %s  ' \
               'base_matrix: %s' % (
                   self.pageid, self.table_no, self.bbox, self.base_matrix
               )
    
    def commandline_check_tabkle(self):
        pass

    def table_to_excel(self, worksheet, start_row=1, start_col=1, **kwargs):
        """write the base matrix then merge and fill in content, all value is 1 or 0
        return a new start row and start col (this may remain the same"""
        # first adjust ranges
        worksheet = adjust_range_size(ws=worksheet, shape=self.base_matrix.shape, start_row=start_row,
                                      start_col=start_col, horizontal_list=self.horizontal_list,
                                      vertical_list=self.vertical_list)

        def cell_to_excel(c):
            cell_matrix = c.virtual_matrix
            position = np.transpose(np.nonzero(cell_matrix))
            logging.info('this is position : %s', position)
            try:
                if 0 < cell_matrix.sum() < 2:
                    # not merge

                    return False, position
                elif cell_matrix.sum() > 1:
                    # merge
                    start = [999, 999]
                    end = [0, 0]
                    for m in position:
                        start = [min(start[0], m[0]), min(start[1], m[1])]
                        end = [max(end[0], m[0]), max(end[1], m[1])]
                    logging.info('this is start %s and end %s'%(start, end))
                    return True, (start, end)
            except AttributeError:
                pass

        logging.info("this is return location row: %s, col: %s"% (self.base_matrix.shape[0] + start_row, start_col))
        for cell in self.doc:
            content = ''
            for c in cell.text:
                content += c

            current_matrix = cell_to_excel(cell)
            logging.info("cell %s 's content is  %s "%(cell.cell_id, content))
            pos = (current_matrix[1][0][0] + start_row, current_matrix[1][0][1] + start_col)

            if current_matrix[0]:

                ref = get_column_letter(current_matrix[1][0][1] + start_col) + str(
                    current_matrix[1][0][0] + start_row) + \
                      ":" + get_column_letter(current_matrix[1][1][1] + start_col) + str(
                    current_matrix[1][1][0] + start_row)
                style_range(ws=worksheet, style=normal, cell_range=ref, alignment=cell_alighment)
            the_cell_in_excel = worksheet.cell(row=pos[0], column=pos[1])
            try:
                the_cell_in_excel.value = content
                the_cell_in_excel.style = normal
            except AttributeError:
                pass
            except ValueError:
                the_cell_in_excel.style = "normal"
        logging.info('Page' + str(kwargs['ref'][self.pageid]))
        if worksheet.title != 'Page' + str(kwargs['ref'][self.pageid]):
            worksheet.title = 'Page' + str(kwargs['ref'][self.pageid])

    @classmethod
    def create_tables(cls, obj, pageid, table_no=0, back_up =None):
        for i in obj:
            print(i)
        class FakeCell:
            def __init__(self, bbox, matrix, no):
                self.no = no
                self.bbox = bbox
                self.matrix = matrix

            def __repr__(self):
                return '<number : %s , bbox : %s , matrix : %s' % (self.no, self.bbox, self.matrix)

        def get_bbox(obj):

            x0, y0, x1, y1 = 9999, 9999, 0, 0
            for c in obj:
                x0, y0, x1, y1 = min(c.bbox[0], x0), min(c.bbox[1], y0), max(c.bbox[2], x1), max(c.bbox[3], y1)
            return x0, y0, x1, y1

        def get_matrix_list(obj, table_bbox):
            x_power = set()
            y_power = set()
            x_power.update([int(table_bbox[0])])
            y_power.update([int(table_bbox[3])])
            for cell in obj:
                x0 = int(cell.bbox[0])
                y0 = int(cell.bbox[1])
                x1 = int(cell.bbox[2])
                y1 = int(cell.bbox[3])
                if not any([approxiamtion(x0, x) for x in list(x_power)]):
                    x_power.update([x0])
                if not any([approxiamtion(x1, x) for x in list(x_power)]):
                    x_power.update([x1])
                if not any([approxiamtion(y0, y) for y in list(y_power)]):
                    y_power.update([y0])
                if not any([approxiamtion(y1, y) for y in list(y_power)]):
                    y_power.update([y1])
            return sorted(list(y_power)), sorted(list(x_power))

        def conduct_base_matrix(l1, l2):
            x = len(l1) - 1
            y = len(l2) - 1
            a = np.zeros(shape=(x, y))
            return a

        def generate_fake_cell(bbox, base_matrix, n, p, no):
            new_matrix = np.zeros(shape=base_matrix.shape)
            new_matrix[n, p] = 1
            return FakeCell(bbox=bbox, matrix=new_matrix, no=no)

        def get_fake_cell(matrix_list, base_matrix):
            x = 0
            while x < len(matrix_list[1]) - 1:
                for j in range(len(matrix_list[0]) - 1):
                    current_bbox = (matrix_list[1][x], matrix_list[0][j], matrix_list[1][x + 1], matrix_list[0][j + 1])
                    n = len(matrix_list[0]) - 2 - j
                    p = x
                    yield generate_fake_cell(bbox=current_bbox, base_matrix=base_matrix, n=n, p=p, no=(n, p))
                x += 1

        sorted_obj = sorted(obj, key=lambda x: x.centre[1])
        bbox = get_bbox(obj=sorted_obj)
        matrix_list = get_matrix_list(obj=sorted_obj, table_bbox=bbox)
        temp = sorted(matrix_list[0], reverse=True)
        vertical_list = [abs(temp[i] - temp[i - 1]) * 1.75 for i in range(1, len(temp))]
        horizontal_list = [(matrix_list[1][i] - matrix_list[1][i - 1]) / 3.2 for i in range(1, len(matrix_list[1]))]
        base_matrix = conduct_base_matrix(matrix_list[0], matrix_list[1])
        pop_cell = get_fake_cell(matrix_list=matrix_list, base_matrix=base_matrix)
        for i in pop_cell:
            found = False
            for j in sorted_obj:
                if inside(j, i):
                    j.virtual_matrix_builder(i.matrix)
                    j.table_no = table_no
                    found = True
            if not found:
               
                if back_up:
                    temp = []
                    while len(back_up)>0:
                        cur = back_up.pop(0)
                        if inside(cur,i):
                            cur_cell = Cell(matrix={'pageid': pageid})
                            cur_cell.virtual_matrix_builder(i.matrix)
                            cur_cell.table_no = table_no
                            cur_cell.bbox = i.bbox
                            cur_cell.text = cur.get_text()
                            print(cur_cell)
                        else:
                            temp.append(cur)
                    back_up = temp
 
        yield Table(bbox=bbox, pageid=pageid, doc=sorted_obj, table_no=table_no,
                    base_matrix=base_matrix, horizontal_list=horizontal_list, vertical_list=vertical_list)

    def table_to_df(self):
        """
        table obj to pandas DataFrame
        :return: DataFrame
        """
        shape = self.base_matrix.shape

        class FakeCell:

            fake_cell_list = []

            def __init__(self, value, matrix, pos):
                self.value = value
                self.matrix = matrix
                self.pos = pos
                FakeCell.fake_cell_list.append(self)

            @classmethod
            def sort_fake_cell_list(cls):
                cls.fake_cell_list.sort(key=lambda x: x.pos[1])

        def generate_fake_cell(shape, n, p, value):
            new_matrix = np.zeros(shape=shape)
            new_matrix[n, p] = 1
            return FakeCell(value=value, matrix=new_matrix, pos=(n, p))

        def split_cell(cell, shape):
            """
            split merged cells, and put merged cell into top left position, created
            fake cells on other positions
            :param cell:
            :return:
            """
            cell_matrix = cell.virtual_matrix
            position = np.transpose(np.nonzero(cell_matrix))
            content = ''
            for c in cell.text:
                content += c
            if cell.virtual_matrix.sum() > 1:
                anchor = position[0]
                generate_fake_cell(shape=shape, n=anchor[0], p=anchor[1], value=content)
                for i in range(1, len(position)):
                    generate_fake_cell(shape=shape, n=position[i][0], p=position[i][1], value='')
            else:
                generate_fake_cell(shape=shape, n=position[0][0], p=position[0][1], value=content)

        for c in self.doc:
            split_cell(cell=c, shape=shape)
        ta = [None] * shape[0]
        FakeCell.sort_fake_cell_list()
        for fc in FakeCell.fake_cell_list:
            row = fc.pos[0]
            logging.info("this is fake cell value: %s "%fc.value)
            if ta[row]:
                ta[row].append(fc.value)
            else:
                ta[row] = []
                ta[row].append(fc.value)
        print(ta)
        result = pd.DataFrame(ta)
        logging.info("page_to df now result is %s \n"%result)
        return result, ta

    @classmethod
    def get_tables(cls, objs, pageid, table_no=0, backup = None):
        """a generator: obj is ltpage
        objs is pagecells[pageid].values"""
        table_no = table_no
        for obj in objs:
            yield cls.create_tables(obj, table_no=table_no, pageid=pageid, back_up=backup)
            table_no += 1
        return


class PageContainer(PDFPage):
    """a page with reference to all class"""

    def __init__(self, doc, pageid, attrs, layouts):
        super().__init__(doc, pageid, attrs)
        self.layouts = layouts

    @classmethod
    def create_page_container(cls, pageid, tables=None, outlines=None, image=None):
        layouts = {
            'tables': tables,
            'outlines': outlines,
            'image': image,
        }
        pass