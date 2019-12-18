import os
import openpyxl
from openpyxl.styles import Border, Side

loc = 'C:/Users/yanjingy/PycharmProjects/hpb_pdfmining/excels/protocols/'


def convert_bytes(num):
    """
    this function will convert bytes to MB.... GB... etc
    """
    for x in ['bytes', 'KB', 'MB', 'GB', 'TB']:
        if num < 1024.0:
            return "%3.1f %s" % (num, x)
        num /= 1024.0


def file_size(file_path):
    """
    this function will return the file size
    """
    if os.path.isfile(file_path):
        file_info = os.stat(file_path)
        return convert_bytes(file_info.st_size)


DELETE_DICT = ['Flammability', 'Physical Characteristics', 'Performance Test','Quality performance']
DELETE_DICT_COUNTRY = ['MX', 'JP', 'India', 'Australia', 'Turkey', 'Japan', 'Mexico', 'IN', 'CN','China']

for i in os.listdir(loc):
    # try:
    wb = openpyxl.load_workbook(loc + i)
    ws = wb.active
    max_row = ws.max_row
    current = 1
    delete_list =[]
    for row in ws.iter_rows(min_row=1):
        if row[1].value:
            try:
                if row[1].value.strip() in DELETE_DICT_COUNTRY or row[4].value.strip() in DELETE_DICT:
                    print(row[1].value, row[4].value)
                    print(current)
                    print(row[1].row)
                    delete_list.append(current)
                else:
                    for cell in row:
                        print("calling")
                        db = Side(style='thin', color='000000')
                        cell.border = Border(left=db, top=db, right=db, bottom=db)
            except AttributeError:
                pass
        else:
            break
        current += 1
    print(sorted(delete_list, reverse=True))
    for j in sorted(delete_list, reverse=True):
        ws.delete_rows(j,1)

    wb.save(loc + 'tt1/' + i)
