import os
import openpyxl
import re
import string
from openpyxl.cell import Cell
import pandas as pd
from tag_keys import COUNTRY_LIST
import logging

BEGIN_DICT = [
    'Requirement Title', 'Country/Region/Industry Std', 'Items', 'Regulation', 'Requirement Title',
    'URL (if applicable)', 'Country/Region',
    'Test method', 'Requirement/Limits', 'Product Scope (Define age range and applicable product types)',
    'Exemptions	Protocol Section',
]


def keyword_modify(a_string, special=False):
    'lower letter, no head or tail whitespace, return string without punctuations'
    a_string = re.sub(r'\n', '', re.sub(r'\t', '', a_string.strip()))
    temp = re.sub('_', '', re.sub('-', '', string.punctuation))
    if special:
        return a_string.translate(str.maketrans('', '', temp))
    else:
        return a_string.translate(str.maketrans('', '', string.punctuation)).strip()


def get_protocol_path(protocol_path):
    presetting_path = 'C:/Users/yanjingy/Documents/work/protocol/TUV protocols/'
    if len(protocol_path) > 5:
        return protocol_path
    else:
        return presetting_path


def get_protocol_storage_path(protocol_store):
    presetting_path = 'C:/Users/yanjingy/PycharmProjects/hpb_pdfmining/excels/protocols/'
    if len(protocol_store) > 5:
        return protocol_store
    else:
        return presetting_path


def pull_from_raw_excel(wb, name, protocol_store):
    ws = wb.active

    def get_begin_row(row):
        keyword_lsit = [i.lower() for i in BEGIN_DICT]
        for c in row:
            if isinstance(c, Cell):
                if c.value:
                    _temp = keyword_modify(str(c.value).lower())
                    if _temp in keyword_lsit:
                        return c.row

    begin_row = 1
    for row in ws.iter_rows():
        k = get_begin_row(row)
        if k:
            begin_row = k + 1
            break
    start = 0
    df = pd.DataFrame(columns=['Requirement Title', 'Country/Region ',
                               'Test method', 'Requirement/Limits', 'Protocol Section'])
    while True:

        if ws['C' + str(begin_row)].value and ws['E' + str(begin_row)].value and ws['F' + str(begin_row)]\
                and ws['G' + str(begin_row)].value and ws['J' + str(begin_row)].value:
            current = [ws['C' + str(begin_row)].value, ws['E' + str(begin_row)].value, ws['F' + str(begin_row)].value,
                       ws['G' + str(begin_row)].value, ws['J' + str(begin_row)].value]
            if current[-1] == 'Labeling and Document Review' or \
                    current[-1] == 'Chemical Analysis' or \
                    current[-1] == 'Labeling and Document Review':
                if current[1] in COUNTRY_LIST:
                    logging.debug("find matched row fro target protocol %s", % current)
                    df.loc[start] = current
        else:
            break
        start += 1
        begin_row += 1
    wb.close()
    return df


def main():
    error = open('excels/protocols/error.txt', 'w')
    protocol_path = input("input path of protocols: ")
    protocol_store = input("input path for storage: ")
    input_path = get_protocol_path(protocol_path)
    output_path = get_protocol_storage_path(protocol_store)
    protocol_list = [i for i in os.listdir(input_path) if i[-4:-1] == 'xls']

    for p in protocol_list:
        try:
            logging.debug("Cleaning Protocol : %s", % p)
            wb = openpyxl.load_workbook(input_path + p, data_only=True)
            df = pull_from_raw_excel(wb=wb, name=p, protocol_store=output_path)
            df.to_excel(output_path + p, sheet_name='protocol', index=False)
            error.close()
        except:
            pass


if __name__ == '__main__':
    main()
