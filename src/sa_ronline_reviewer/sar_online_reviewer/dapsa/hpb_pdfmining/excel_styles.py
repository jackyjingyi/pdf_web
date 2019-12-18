from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

# highlight
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=20)

STYLES = {
    'highlight': {
        'font': {'bold': True, 'size': 20, },
        'border': {'style': 'thick', 'color': "000000"},
        'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True}
    },
    'normal': {
        'font': {'bold': False, 'size': 14},
        'border': {'style': 'thin', 'color': '000000'},
        'alignment': {'horizontal': 'general', 'vertical': 'top', 'wrap_text': True},
    },
}

def get_style(*args, **kwargs):
    name = NamedStyle(name=args[0])
    name.font = Font(bold=kwargs['style_dict'][args[0]]['font']['bold'])
    db = Side(style=kwargs['style_dict'][args[0]]['border']['style'],
              color=kwargs['style_dict'][args[0]]['border']['color'])
    name.border = Border(left=db, top=db, right=db, bottom=db)
    name.alignment = Alignment(horizontal=kwargs['style_dict'][args[0]]['alignment']['horizontal'],
                               vertical=kwargs['style_dict'][args[0]]['alignment']['vertical'],
                               wrap_text=kwargs['style_dict'][args[0]]['alignment']['wrap_text'])
    return name


def style_range(ws, style, cell_range, fill=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.
    :param ws: Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """
    border = style.border
    font = style.font
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
    first_cell.alignment = alignment
    rows = ws[cell_range]
    if font:
        first_cell.font = font
    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom
    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def adjust_range_size(ws, shape, horizontal_list, vertical_list, start_row=1, start_col=1, ref=None):
    """
    :param ws: worksheet
    :param shape: matrix shape , represent a table
    :param horizontal_list: a list to adjust col width [w1,w2]
    :param vertical_list: a list to adjust row height [h1,h2]
    :param start_row: start from row index  int
    :param start_col: start from column index  int
    :param ref: eg."A3:F20"
    :return: worksheet with adjusted range
    """
    if ref:
        # with ref then, shape , start_row ,start_col is unnecessary, actually conflict
        pass
    else:
        c = start_col
        while len(horizontal_list) > 0:
            current_width = horizontal_list.pop(0)
            ws.column_dimensions[get_column_letter(c)].width = current_width
            c += 1
        r = start_row
        while len(vertical_list) > 0:
            current_height = vertical_list.pop(0)
            ws.row_dimensions[r].height = current_height
            r += 1
        assert c == shape[1] + start_col, "column added {} but should add {}".format(c - start_col, shape[1])
        assert r == shape[0] + start_row, "row added {} but should add {}".format(r - start_row, shape[0])
    return ws


