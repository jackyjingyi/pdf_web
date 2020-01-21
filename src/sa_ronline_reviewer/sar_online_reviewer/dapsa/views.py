import re
import os
import time
import logging
import pandas as pd
import string
import pdfminer
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFTextExtractionNotAllowed
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.layout import LAParams, LTChar, LTCurve, LTTextLine, LTText, LTTextBox, LTLine, LTRect, LTImage, \
    LTTextBoxHorizontal, LTPage
from pdfminer.converter import PDFPageAggregator
from pdfminer.pdfpage import PDFPage
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import HttpResponse
from django.views import View
from .models import Protocols
from .hpb_pdfmining.pdf_layouts import Point, Cluster, Cell, Table, approxiamtion, inside
from .hpb_pdfmining.tag_keys import SKIP_KEYS, BEGIN_KEYS, INFO_KEYS, DATE_PATTERN, MONTH
from .hpb_pdfmining.mapper import predefine_, pdf_df_rename, get_protocol, Mapper
from .hpb_pdfmining.analyze_class import check_dict, keyword_modify, keyword_contains_in, resove_dict
from .hpb_pdfmining.main_decorators import timeout, check_first_page_with_dict, check_not_first_page_with_dict, controller
from .hpb_pdfmining.extractors import FirstPageInfo, user_select_from_list, PDFExtractor, PageExtractor, checkpointsvaliadation, append_df_to_excel, find_key, find_right_neightbors, PDFExtractorFitz, PageExtractorFitz, Searcher
from .models import MiningLog, TestDoc, SamplePool, Protocols, Protocol, AsinAssign, HubbleInfo,MiningQueue
from .forms import TestDocForm
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import load_workbook
import uuid
"""
class to render or serilize the output to json and then pass it to frontend
1. input :  pdf address reference
2. 
"""
today_min = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
today_max = datetime.datetime.combine(datetime.date.today(), datetime.time.max)


def home(request):
    count = User.objects.count()
    return render(request, 'home.html', {
        'count': count
    })


def signup(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = UserCreationForm()
    return render(request, 'registration/signup.html', {
        'form': form
    })


@login_required
def secret_page(request, *args, **kwargs):
    _temp = {
        'SPB': '1',
        'HPB': '2',
        'CPB': '3',
    }
    task_type = kwargs['task_type']
    user = request.user
    # check if
    SamplePool.assign_asin(number=8, task_type=task_type)
    asins = AsinAssign.objects.filter(
        assignee=request.user, assign_date__range=(today_min, today_max), task_type=_temp[task_type])
    print(asins)
    if len(asins) == 0:
        # new user
        SamplePool.assign_single(number=8, user=user, task_type=task_type)
        asins = AsinAssign.objects.filter(
            assignee=request.user, assign_date__range=(today_min, today_max), task_type=_temp[task_type])
    asins = [i.asin for i in asins]

    return render(request, 'secret_page.html', {'asins': asins, 'task_type': task_type})


class SecretPage(LoginRequiredMixin, TemplateView):
    template_name = 'secret_page.html'


def tasks(request, asin_number, task_type, *args, **kwargs):
    # Handle file upload
    asin_number = asin_number
    task_type = task_type
    print(asin_number)
    print(task_type)
    
    if request.method == 'POST':

        form = TestDocForm(request.POST, request.FILES)
        if form.is_valid():
            temp_caseid = str(uuid.uuid4())

            newdoc = TestDoc(caseid=temp_caseid, pl=request.POST['pl'],
                             docfile=request.FILES['docfile'], uploaded_by=request.user, asin=asin_number)  # .save()

            newdoc.save()

            # excel_file = request.FILES['docfile']
            # wb = load_workbook(excel_file)
            # ws = wb.worksheets[0]
            # f = open('log.txt', 'w')
            # for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            #     try:
            #         HubbleInfo(asin = row[0].value, mfp = row[1].value, mfp_code = row[2].value, approval_type = row[3].value,
            #         life_cycle =row[4].value, description = row[5].value,date_approved=row[6].value, experiration_date = row[7].value,
            #         active_region = row[8].value,regular_body=row[9].value,po_approve=row[10].value,
            #         first_request_date =row[11].value, approver=row[12].value, escalation_tt = row[13].value,
            #         compliance_note = row[14].value).save()
            #     except:
            #         f.writelines([str(i.value)+',' for i in row]+['\n'])
            #Protocols(protocol_name=row[0].value, short_cut=row[1].value, amazon_number=row[3].value,uploaded_by = request.user).save()

            # SamplePool(parent_asin=row[0].value, region_id=row[1].value, marketplace_id=row[2].value, asin=row[3].value,
            #           ifdone=row[4].value, brand_code=row[5].value, item_name=row[6].value, style=row[7].value, risk_level=row[8].value,
            #           classification_type=row[9].value, product_type=row[10].value, brand_name=row[
            #               11].value, product_group_description=row[12].value,
            #           parent_asin_name=row[13].value, edition=row[14].value,
            #           max_available_quantity=row[15].value, min_available_quantity=row[
            #               16].value, eod_available_quantity=row[17].value,
            #           all_vendors=row[18].value, first_vendor=row[19].value, vendor_name=row[
            #               20].value, sub_parent_asin_number=row[21].value,
            #           ifpick=row[22].value, assign_status='3').save()

            # Redirect to the document list after POST
            # try:
            #    current_protocol = Protocols.objects.get(
            #        protocol_name=row[1].value)
            #    idx = row[0].value
            #    if idx > len(Protocol.objects.all()):
            #        Protocol(protocol_name=current_protocol, speck_number=row[2].value, regulation=row[3].value,
            #                 requirement_title=row[4].value, link=row[5].value, region=row[6].value,  test_method=row[7].value,
            #                 requirement=row[8].value, protduct_scope=row[9].value, exemption=row[
            #                     10].value, protocol_section=row[11].value,
            #                 mandatory_voluntary=row[12].value, reationale=uuid.uuid4(), inner_id=row[13].value, is_cornerstone=True, uploaded_by_1=request.user, last_updated_by_1=request.user).save()
            #        print("saving")
            #    else:
            #        print("not saving")
            #    print(len(Protocol.objects.all()))
            # except:
            #    f.writelines([str(i.value)+',' for i in row]+['\n'])
            return redirect(reverse('tasks', args=[task_type, asin_number]))
    else:
        form = TestDocForm()  # A empty, unbound form

    # Load documents for the list page
    documents = TestDoc.objects.filter(asin=asin_number)

    asin = SamplePool.objects.get(asin=asin_number)
    hubbles= HubbleInfo.objects.filter(asin = "B001TOD7ME")
    print(hubbles)
    # Render list page with the documents and the form
    return render(request, 'dapsa/tasks.html', {'asin_number':asin_number, 'task_type':task_type,'documents': documents, 'asin': asin, 'form': form, 'hubbles':hubbles})

def create_queue(casid):
    pass    


def pdfmain(request, asin_number, task_type, caseid,*args, **kargs):
    asin_number = asin_number
    task_type =task_type
    caseid = caseid
    # t = MiningQueue.objects.get(sessionid = caseid)
    document = TestDoc.objects.get(caseid=caseid)
    if document:
        results = pdf_analysis_main(document)
        if isinstance(results, set) and len(results) > 0:
            current_protocol = [w[4] for w in results]
            print(current_protocol)
            pool = [str(i.protocol_name).lower()
                    for i in list(Protocols.objects.all())]
            print(pool)
            t = [current_protocol[0].lower() in p for p in pool]
            idx = int(re.findall(r'\d+', current_protocol[0].lower())[0])
            print(idx)
            if isinstance(idx, int):

                p_obj = Protocols.objects.filter(
                    amazon_number=idx)[0]
                # corresponding protocol items
                print(p_obj.protocol_name)
                protcol_item = Protocol.objects.filter(
                    protocol_name_id=p_obj)
            else:
                # for check
                p_obj = Protocols.objects.get(amazon_number=21)
                protcol_item = Protocol.objects.filter(
                    protocol_name_id=p_obj)
            # protocol report used
            results = {}
            results['protocol'] = p_obj.protocol_name
            conclusions = conclusion_extraction(
                doc=document, protocol=p_obj.id, begin=4, end=10)
            (conclusion_dict, homeless_items, p_set) = mapping_protocols(
                protcocol_set=[i for i in protcol_item], extract_list=conclusions)
            print(conclusions)
            all_protocols = [i.protocol_name for i in Protocols.objects.all()]
            if p_obj:
                k = all_protocols.index(p_obj.protocol_name)
                del all_protocols[k]
                all_protocols = [p_obj.protocol_name] + all_protocols
            return render(request, 'dapsa/analysis.html', {'asin_number':asin_number,'task_type':task_type,'caseid':caseid,'document': document, 'dict': results, 'conclusions': conclusion_dict, 'p_name': p_obj, 'miss': p_set,'all_protocols':all_protocols})
        # return render(request, 'dapsa/analysis.html', {'document': document, 'dict': results})
    return render(request, 'dapsa/analysis.html', {'asin_number':asin_number,'task_type':task_type,'document': document})


def conclusion_extraction(*args, **kwargs):
    """
        **kwargs :
            doc => filefield
            protocol => int : id of protocols 
            begin => int: page to start analysis conclusion tables
            end => int: page number to stop analysis
            sessionid =>uuid for pdf review

    """
    Point.purge()
    Cluster.purge()
    Cell.purge()
    doc = kwargs['doc']
    protocol = kwargs['protocol']
    begin = kwargs['begin']
    end = kwargs['end']
    fp = doc.docfile.open('rb')
    pdf = PDFExtractor(fp)

    # testing mode set start page
    pdf.begin_page = begin
    # start from 'start' page, pdf_iter is a generator, generate "PageExtractor" instance
    pdf_iter = pdf.process_page()
    # testing : only process 9 pages start from begin page
    count = 0
    conclusion_table = []
    while count < end - begin:
        page = next(pdf_iter)
        # PageExtractor call() {
        #   1.detect points from ltrect or ltline
        #   2.clustering points into Cluster objects
        #   3.establish cells from clusters
        #   4.establish tables by cells
        # }
        page()
        # To get conclusion table, the perviouse move is get the last table object
        # in a page, however, we can transform it into if table in fitz.Rect(page.body)
        count += 1
        print(page.tables)
        for table in page.tables:
            # df is a pd.dataframe, first column should be amazon speck number
            # because cornerstone protocols are simplier than TUV protocols,
            # new mapping function can simply mapping with spec number to identifiy a record

            df = table.table_to_df()
            print(df)
        # 1 is list which is the source of df [[str,str], []]
        conclusion_table += [i + [page.page_number] for i in page.tables[-1].table_to_df()[1]]
    # pdf purge

    return conclusion_table


def mapping_protocols(protcocol_set, extract_list, *args, **kwargs):
    """
    no validation 
    1. extract list possibly coming in page unit

    """
    print(extract_list)
    def check_outcome(p, f, c):
        if "pass" in p.lower():
            return "pass"
        elif "fail" in f.lower():
            return "fail"
        else:
            return c

    try:
        assert len(protcocol_set) > 0, "empty protocol list"
        assert len(extract_list) > 0, "empty extract list"
    except AssertionError:
        logging.warning("one of mapping list is empty")
        return

    results = {
        'PASS': [],
        'FAIL': [],
        'NR': []
    }
    homeless_items = []
    while len(extract_list) > 0:
        _current_item = extract_list.pop(0)
        if len(protcocol_set) == 0:
            break
        try:
            _current_req = _current_item[0].strip()
        except TypeError:
            continue
        _pass = _current_item[-4].strip()
        _fail = _current_item[-3].strip()
        _comment = _current_item[-2].strip()
        print(_current_req, _pass, _fail, _comment)
        if len(_current_req) > 0:
            _end_point = check_outcome(_pass, _fail, _comment)
            _re_item = re.findall(r'^[a-zA-Z]\d+', _current_req)
            print(_re_item)
            try:
                #assert isinstance(_re_item, list), "Not a list"
                assert any([i == _current_req for i in _re_item]
                           ), "it might be No Spec"
                # if with a spec number jump in to a for loop
                for idx, p in enumerate(protcocol_set):
                    # exact check
                    if p.speck_number.lower() == _current_req.lower():
                        if _end_point == "pass":
                            results['PASS'].append((p, _current_item))
                            del protcocol_set[idx]
                            break
                        elif _end_point == "fail":
                            results['FAIL'].append((p, _current_item))
                            del protcocol_set[idx]
                            break
                        else:
                            results['NR'].append((p, _current_item))
                            del protcocol_set[idx]
                            break
            except AssertionError:
                # No Spec extracted items
                logging.info("No spec pass it to homeless")
                homeless_items.append(_current_item)
        else:
            continue

    return (results, homeless_items, protcocol_set)


def conclusion_check(request, asin_number,task_type,caseid,speck,*args, **kwargs):
    
    conclusions = request.GET.get("conclusions")
    print(conclusions)
    return render(request, 'dapsa/check.html',{'asin_number':asin_number,'task_type':task_type,'caseid':caseid,'speck':speck})


def secret_page_protocol_audit(request):
    protocols = Protocols.objects.all()

    protocol_detail = Protocol.objects.filter(id__lt=10)
    print(protocol_detail)
    return render(request, 'secret_page_protocol.html', {'protocols': protocols, 'protocol_detail': protocol_detail})

    # return render(request, 'secret_page_protocol.html', {'protocols': protocols})


def pdf_analysis_main(doc, begin = 0, end=5):
    # frist create an instance to Mining Queue, if there are workers avaliable, pop queue,
    # delete this records, (or create a status, change that status to mining)
    # => analyse pdf and change the MiningLog status continue to monitoring this record
    # step1 finish => display bascinfo and require users' confirmation
    # start step2, display chosen protocol
    # analyzing and mapping , change the mapped protocol records dynamically
    #
    def _find_keyword(page, s, words):
        search_result = s.page_search_for(page=page,words = words)
        return search_result

    doc = PDFExtractorFitz(doc.docfile.path)
    page = doc.get_page(0)
    # i = 0 
    # while i <= end - begin:
    #     page = doc.get_page(begin + i)
    #     page.body = 0.3
    #     page.header = 0.2
    #     p_list = page.page.getTextWords()
    #     p_block = page.page.getText("blocks")
    #     p_list_header = page.recover(words=p_list, rect=page.header)
    
    p_list = page.page.getTextWords()
    p_block = page.page.getText("blocks")
    page.body = 0.3
    page.header = 0.2
    p_list_block_header = page.check_blocks(words=p_block, rect=page.header)
    # for item in p_list_block_header:
    #     print("header information:~~")
    #     print(item)
    
    searcher = Searcher(idx=0, key="protocol")
    results = searcher.page_search_for(flag=1, page=page, )
    new_searcher = Searcher(idx=0, key="report number")
    new_results = new_searcher.page_search_for(page = page, flag = 1,words = p_list_block_header)

    print(results, type(results))
    print(new_results,type(new_searcher))
    # assert 1 ==2
    return results


class PDFExtractionView(View):

    def get(self, request, *args, **kwargs):
        return HttpResponse('Hello world')
